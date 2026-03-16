"""Export generation: Excel decision table and NL text statements.

Implements PRD F-24 through F-27.
Export is gated by unresolved flags — the API layer enforces this.
"""

import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from schema import ExtractedRule, RuleStatus


# ── Excel Export (F-24, F-26) ──

EXCEL_HEADERS = [
    "Rule ID", "Category", "Field", "Operator", "Value", "Unit",
    "Conditions", "Condition Logic", "Outcome", "Failure Outcome",
    "NL Statement", "Source Quote", "Source Section", "Source Page",
    "Precedence", "Rule Scope", "Overrides Rule", "Footnote",
    "Canonical Field", "Canonical Category", "Status",
]

STATUS_COLORS = {
    "approved": "C6EFCE",       # Green
    "rejected": "FFC7CE",       # Red
    "flagged_regulatory": "FFEB9C",  # Yellow
    "flagged_uncertain": "FFEB9C",   # Yellow
    "pending_review": "D9E1F2",      # Light blue
}


def generate_excel(rules: list[ExtractedRule], doc_name: str) -> io.BytesIO:
    """Generate formatted Excel decision table from approved rules.

    Args:
        rules: List of rules to export (should be filtered to approved only)
        doc_name: Source document name for the sheet header

    Returns:
        BytesIO buffer containing the .xlsx file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Decision Table"

    # Title row
    ws.merge_cells("A1:U1")
    title_cell = ws["A1"]
    title_cell.value = f"Mortgage Lending Rules — Extracted from {doc_name}"
    title_cell.font = Font(size=14, bold=True, color="1F4E79")
    title_cell.alignment = Alignment(horizontal="center")

    # Metadata row
    ws.merge_cells("A2:U2")
    meta_cell = ws["A2"]
    meta_cell.value = (
        f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')} | "
        f"Total Rules: {len(rules)} | "
        f"EU AI Act: HIGH RISK — All rules require human review"
    )
    meta_cell.font = Font(size=10, italic=True, color="666666")

    # Header row (row 4)
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col, header in enumerate(EXCEL_HEADERS, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border

    # Data rows
    for row_idx, rule in enumerate(rules, 5):
        status_val = rule.status.value if hasattr(rule.status, "value") else str(rule.status)
        category_val = rule.category.value if hasattr(rule.category, "value") else str(rule.category)
        scope_val = rule.rule_scope.value if rule.rule_scope and hasattr(rule.rule_scope, "value") else str(rule.rule_scope or "")

        values = [
            rule.rule_id,
            category_val,
            rule.field,
            rule.operator,
            rule.value,
            rule.unit or "",
            _format_conditions(rule.conditions),
            rule.condition_logic or "",
            rule.outcome,
            rule.failure_outcome or "",
            rule.nl_statement,
            rule.source_quote,
            rule.source_section or "",
            rule.source_page or "",
            rule.precedence or "",
            scope_val,
            rule.overrides_rule_id or "",
            rule.footnote_ref or "",
            rule.canonical_field or "",
            rule.canonical_category or "",
            status_val,
        ]

        fill_color = STATUS_COLORS.get(status_val, "FFFFFF")
        row_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.fill = row_fill
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Auto-width columns (capped at 50)
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

    # Freeze header row
    ws.freeze_panes = "A5"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# ── NL Text Export (F-25) ──

def generate_nl_text(rules: list[ExtractedRule], doc_name: str) -> str:
    """Generate natural language rule statements grouped by category.

    Args:
        rules: List of rules to export
        doc_name: Source document name

    Returns:
        Formatted text string
    """
    lines = [
        "MORTGAGE LENDING RULES — NATURAL LANGUAGE SUMMARY",
        "=" * 55,
        f"Source: {doc_name}",
        f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
        f"Total Rules: {len(rules)}",
        "",
        "EU AI Act Classification: HIGH RISK",
        "All rules require human review and compliance sign-off before deployment.",
        "",
    ]

    # Group by category
    by_category: dict[str, list[ExtractedRule]] = {}
    for rule in rules:
        cat = rule.category.value if hasattr(rule.category, "value") else str(rule.category)
        by_category.setdefault(cat.upper(), []).append(rule)

    for cat in sorted(by_category):
        lines.append(f"\n## {cat}")
        lines.append("-" * 40)
        for rule in by_category[cat]:
            flag_marker = ""
            if rule.guardrail_flags:
                flag_marker = " [FLAGGED]"
            lines.append(f"  {rule.rule_id}: {rule.nl_statement}{flag_marker}")
            if rule.conditions:
                lines.append(f"    Conditions: {_format_conditions(rule.conditions)}")
            if rule.source_section:
                lines.append(f"    Source: {rule.source_section}, p.{rule.source_page or '?'}")
        lines.append("")

    return "\n".join(lines)


# ── Export Gating (F-17, F-27) ──

def can_export(rules: list[ExtractedRule]) -> tuple[bool, list[str]]:
    """Check if export is allowed — blocked by unresolved flags.

    Returns:
        Tuple of (can_export, list of blocking reasons)
    """
    blockers = []
    for rule in rules:
        status = rule.status.value if hasattr(rule.status, "value") else str(rule.status)
        if status == "flagged_regulatory":
            blockers.append(f"{rule.rule_id}: unresolved regulatory flag")
        elif status == "flagged_uncertain":
            blockers.append(f"{rule.rule_id}: unresolved uncertainty flag — manual logic required")
    return (len(blockers) == 0, blockers)


def _format_conditions(conditions: dict | list | None) -> str:
    """Format conditions for display."""
    if not conditions:
        return ""
    if isinstance(conditions, dict):
        return "; ".join(f"{k}={v}" for k, v in conditions.items())
    if isinstance(conditions, list):
        return "; ".join(str(c) for c in conditions)
    return str(conditions)
